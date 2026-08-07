"""Nguồn dữ liệu ngoài cho bot tra cứu (Điều 6: "Có khả năng kết nối đến cơ sở
dữ liệu (ví dụ bảng excel ... để truy cứu thông tin và tư vấn cho KH)").

HAI CHẾ ĐỘ, cho hai nhu cầu khác hẳn nhau — đừng gộp làm một:

  mode="rag"     TRI THỨC CHUNG. Bảng giá, biểu phí, danh mục sản phẩm. Mỗi dòng
                 thành một mảnh văn bản, nạp vào ChromaDB, tìm gần đúng theo
                 nghĩa. Hỏi "lãi suất vay mua nhà" thì tìm ra dòng nói về nó.

  mode="lookup"  TRA THEO KHÁCH. "Dư nợ của tôi bao nhiêu", "đơn hàng tới đâu".
                 Đây là tra CHÍNH XÁC theo số điện thoại / mã khách, không phải
                 tìm gần đúng. Nhét vào RAG là sai bản chất: nó sẽ trả về dòng
                 của một khách khác có nội dung "giống" hơn.

RÀNG BUỘC OFFLINE: chỉ nhận file cục bộ và CSDL trong mạng nội bộ. Đường dẫn
http(s) bị từ chối thẳng — hợp đồng ghi "Chạy offline 100%, ko call API qua các
web khác".
"""

import asyncio
import csv
import json
import logging
import re
import sqlite3
import time
import uuid
from pathlib import Path

from backend.models import db

logger = logging.getLogger(__name__)

KINDS = ("excel", "csv", "sqlite")
MODES = ("rag", "lookup")

# Trần số dòng nạp một lần. Bảng vài trăm nghìn dòng nạp thẳng vào ChromaDB sẽ
# ăn hết RAM và làm treo cả tiến trình đang phục vụ cuộc gọi.
_TRAN_DONG = 20_000


def _row(r) -> dict:
    d = dict(r)
    try:
        d["config"] = json.loads(d.get("config") or "{}")
    except (ValueError, TypeError):
        d["config"] = {}
    d["enabled"] = bool(d.get("enabled"))
    return d


# --- kho ----------------------------------------------------------------------

def _list_sync() -> list[dict]:
    conn = db.connection()
    if conn is None:
        return []
    with db.write_lock:
        rows = conn.execute("SELECT * FROM data_sources ORDER BY created_at").fetchall()
    return [_row(r) for r in rows]


async def list_sources() -> list[dict]:
    return await asyncio.to_thread(_list_sync)


def _get_sync(source_id: str) -> dict | None:
    conn = db.connection()
    if conn is None or not source_id:
        return None
    with db.write_lock:
        row = conn.execute(
            "SELECT * FROM data_sources WHERE source_id = ?", (source_id,)
        ).fetchone()
    return _row(row) if row else None


async def get_source(source_id: str) -> dict | None:
    return await asyncio.to_thread(_get_sync, source_id)


def _upsert_sync(source_id: str, fields: dict) -> dict | None:
    conn = db.connection()
    if conn is None:
        return None
    if not source_id:
        source_id = f"ds_{uuid.uuid4().hex[:8]}"

    data = {
        "name": fields.get("name", "Nguồn dữ liệu"),
        "kind": fields.get("kind") if fields.get("kind") in KINDS else "excel",
        "config": json.dumps(fields.get("config") or {}, ensure_ascii=False),
        "mode": fields.get("mode") if fields.get("mode") in MODES else "rag",
        "lookup_key": fields.get("lookup_key", ""),
        "knowledge_tag": fields.get("knowledge_tag", ""),
        "enabled": 1 if fields.get("enabled", True) else 0,
    }
    cols = ", ".join(data)
    with db.write_lock, conn:
        conn.execute(
            f"INSERT INTO data_sources (source_id, {cols}, created_at) "
            f"VALUES (?, {', '.join('?' * len(data))}, ?) "
            f"ON CONFLICT(source_id) DO UPDATE SET "
            + ", ".join(f"{k} = excluded.{k}" for k in data),
            (source_id, *data.values(), time.time()),
        )
    return _get_sync(source_id)


async def upsert_source(source_id: str = "", **fields) -> dict | None:
    return await asyncio.to_thread(_upsert_sync, source_id, fields)


def _delete_sync(source_id: str) -> bool:
    conn = db.connection()
    if conn is None:
        return False
    with db.write_lock, conn:
        cur = conn.execute("DELETE FROM data_sources WHERE source_id = ?", (source_id,))
    return cur.rowcount > 0


async def delete_source(source_id: str, rag=None) -> bool:
    """Xoá nguồn. Nếu nó đã nạp vào RAG thì gỡ luôn các mảnh của nó ra."""
    nguon = await get_source(source_id)
    if nguon is None:
        return False
    if rag is not None and nguon["mode"] == "rag":
        await asyncio.to_thread(rag.xoa_theo_nguon, _ma_nguon(source_id))
    return await asyncio.to_thread(_delete_sync, source_id)


def _ghi_ket_qua_sync(source_id: str, so_dong: int, loi: str):
    conn = db.connection()
    if conn is None:
        return
    with db.write_lock, conn:
        conn.execute(
            "UPDATE data_sources SET row_count = ?, sync_at = ?, last_error = ? "
            "WHERE source_id = ?",
            (so_dong, time.time(), loi, source_id),
        )


def _ma_nguon(source_id: str) -> str:
    """Nhãn `source` gắn vào metadata RAG, để xoá/nạp lại theo từng nguồn."""
    return f"data_source:{source_id}"


# --- đọc dữ liệu --------------------------------------------------------------

def _kiem_duong_dan(raw: str) -> tuple[Path | None, str]:
    """Chỉ chấp nhận file cục bộ. Trả về (đường dẫn, lỗi)."""
    duong_dan = (raw or "").strip()
    if not duong_dan:
        return None, "Chưa khai đường dẫn file"
    if duong_dan.lower().startswith(("http://", "https://", "ftp://")):
        return None, (
            "Hệ thống chạy offline 100% theo hợp đồng nên không nhận đường dẫn "
            "web — tải file về máy rồi khai đường dẫn cục bộ"
        )
    p = Path(duong_dan).expanduser()
    if not p.exists():
        return None, f"Không tìm thấy file: {p}"
    if not p.is_file():
        return None, f"Đường dẫn không phải file: {p}"
    return p, ""


def _doc_excel(p: Path, sheet: str = "") -> tuple[list[dict], str]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return [], "Chưa cài openpyxl — chạy: pip install openpyxl"

    try:
        # read_only + data_only: file bảng giá vài chục nghìn dòng mà nạp cả
        # công thức lẫn định dạng thì tốn RAM gấp nhiều lần, mà ta chỉ cần giá trị.
        wb = load_workbook(str(p), read_only=True, data_only=True)
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
        hang = ws.iter_rows(values_only=True)
        tieu_de = next(hang, None)
        if not tieu_de:
            return [], "File rỗng"
        cot = [str(c).strip() if c is not None else f"cot_{i}"
               for i, c in enumerate(tieu_de, 1)]

        rows = []
        for r in hang:
            if r is None or all(c is None or str(c).strip() == "" for c in r):
                continue
            rows.append({
                cot[i]: _bo_nhay(v) for i, v in enumerate(r) if i < len(cot)
            })
            if len(rows) >= _TRAN_DONG:
                break
        wb.close()
        return rows, ""
    except Exception as e:
        return [], f"Không đọc được file Excel: {e}"


def _doc_csv(p: Path) -> tuple[list[dict], str]:
    try:
        with p.open("r", encoding="utf-8-sig", errors="replace", newline="") as f:
            reader = csv.DictReader(f)
            rows = []
            for r in reader:
                rows.append({(k or "").strip(): _bo_nhay(v) for k, v in r.items()})
                if len(rows) >= _TRAN_DONG:
                    break
        return rows, ""
    except Exception as e:
        return [], f"Không đọc được file CSV: {e}"


def _doc_sqlite(p: Path, bang: str, cau_lenh: str = "") -> tuple[list[dict], str]:
    try:
        conn = sqlite3.connect(f"file:{p}?mode=ro", uri=True)
        conn.row_factory = sqlite3.Row
        try:
            if cau_lenh:
                # Chỉ cho SELECT: đây là nguồn ĐỌC, không phải nơi chạy lệnh tuỳ ý.
                if not cau_lenh.strip().lower().startswith("select"):
                    return [], "Chỉ chấp nhận câu lệnh SELECT"
                cur = conn.execute(cau_lenh)
            elif bang:
                if not bang.replace("_", "").isalnum():
                    return [], f"Tên bảng không hợp lệ: {bang}"
                cur = conn.execute(f"SELECT * FROM {bang} LIMIT {_TRAN_DONG}")
            else:
                return [], "Chưa khai tên bảng hoặc câu lệnh SELECT"
            rows = [{k: _bo_nhay(v) for k, v in dict(r).items()}
                    for r in cur.fetchmany(_TRAN_DONG)]
            return rows, ""
        finally:
            conn.close()
    except Exception as e:
        return [], f"Không đọc được SQLite: {e}"


async def doc_du_lieu(nguon: dict) -> tuple[list[dict], str]:
    """Đọc toàn bộ dòng của một nguồn. Trả về (dòng, lỗi)."""
    cfg = nguon.get("config") or {}
    kind = nguon.get("kind", "excel")

    if kind == "sqlite":
        p, loi = _kiem_duong_dan(cfg.get("path", ""))
        if loi:
            return [], loi
        return await asyncio.to_thread(
            _doc_sqlite, p, cfg.get("table", ""), cfg.get("query", "")
        )

    p, loi = _kiem_duong_dan(cfg.get("path", ""))
    if loi:
        return [], loi
    if kind == "csv":
        return await asyncio.to_thread(_doc_csv, p)
    return await asyncio.to_thread(_doc_excel, p, cfg.get("sheet", ""))


# --- chế độ rag: nạp vào ChromaDB ---------------------------------------------

def _dong_thanh_van_ban(row: dict, cot_noi_dung: list[str]) -> str:
    """Một dòng bảng -> một câu tiếng Việt đọc được.

    "Sản phẩm: Vay tín chấp | Lãi suất: 7.9%/năm | Hạn mức: 500 triệu"

    Giữ nguyên tên cột trong chuỗi: mô hình cần biết con số đó LÀ GÌ, chứ một
    dãy giá trị trần trụi thì nó sẽ gán bừa.
    """
    phan = []
    for k, v in row.items():
        if cot_noi_dung and k not in cot_noi_dung:
            continue
        v = (v or "").strip()
        if v:
            phan.append(f"{k}: {v}")
    return " | ".join(phan)


async def nap_vao_rag(nguon: dict, rag) -> dict:
    """Đọc nguồn rồi nạp vào ChromaDB. Xoá bản cũ của chính nguồn này trước."""
    rows, loi = await doc_du_lieu(nguon)
    if loi:
        await asyncio.to_thread(_ghi_ket_qua_sync, nguon["source_id"], 0, loi)
        return {"error": loi}
    if not rows:
        return {"error": "Nguồn không có dòng dữ liệu nào"}

    ma = _ma_nguon(nguon["source_id"])
    cot_noi_dung = (nguon.get("config") or {}).get("content_columns") or []

    def _nap():
        da_xoa = rag.xoa_theo_nguon(ma)
        van_ban = [_dong_thanh_van_ban(r, cot_noi_dung) for r in rows]
        van_ban = [t for t in van_ban if t]
        if not van_ban:
            return da_xoa, 0
        # Gom nhiều dòng vào một tài liệu: mỗi dòng một mảnh riêng thì top_k=3
        # chỉ nhìn được 3 dòng của cả bảng giá, gần như luôn thiếu.
        khoi, buoc = [], 12
        for i in range(0, len(van_ban), buoc):
            khoi.append("\n".join(van_ban[i:i + buoc]))
        for i, noi_dung in enumerate(khoi):
            rag.ingest_text(
                noi_dung,
                doc_id=f"{nguon['source_id']}_{i}",
                metadata={
                    "source": ma,
                    "ten_nguon": nguon.get("name", ""),
                    "tag": nguon.get("knowledge_tag", ""),
                },
            )
        return da_xoa, len(khoi)

    try:
        da_xoa, so_khoi = await asyncio.to_thread(_nap)
    except Exception as e:
        loi = f"Nạp vào RAG hỏng: {e}"
        logger.warning(loi, exc_info=True)
        await asyncio.to_thread(_ghi_ket_qua_sync, nguon["source_id"], 0, loi)
        return {"error": loi}

    await asyncio.to_thread(_ghi_ket_qua_sync, nguon["source_id"], len(rows), "")
    logger.info(
        f"[nguồn dữ liệu] {nguon['name']}: nạp {len(rows)} dòng "
        f"({so_khoi} khối), xoá {da_xoa} mảnh cũ"
    )
    return {"so_dong": len(rows), "so_khoi": so_khoi, "da_xoa_cu": da_xoa}


# --- chế độ lookup: tra chính xác theo khoá -----------------------------------

_CHI_SO = re.compile(r"^\+?\d[\d\s.\-()]*$")


def _chuan_khoa(s: str) -> str:
    """Chuẩn hoá khoá trước khi so khớp.

    Số điện thoại trong Excel viết đủ kiểu: '0912 345 678', '+84912345678',
    '0084912345678', '0912.345.678'. Cả bốn là MỘT thuê bao, nên phải quy về
    cùng một dạng chứ không chỉ bỏ ký tự lạ — bỏ ký tự lạ thôi thì '+84912345678'
    thành '84912345678', vẫn khác '0912345678' và tra không ra.

    Dùng lại đúng `contacts_db.normalize_phone` để danh bạ và nguồn dữ liệu
    ngoài hiểu số giống hệt nhau.
    """
    raw = _bo_nhay(s)
    if _CHI_SO.match(raw):
        from backend.models.contacts_db import normalize_phone
        chuan = normalize_phone(raw)
        if len(chuan) >= 8:
            return chuan
    return "".join(c for c in raw.lower() if c.isalnum())


def _bo_nhay(v) -> str:
    """Bỏ dấu nháy đơn Excel chèn vào đầu ô để ép kiểu chữ.

    Không bỏ thì số điện thoại hiện ra trong prompt là "'0912 345 678" và mô
    hình đọc cả dấu nháy cho khách nghe.
    """
    return str("" if v is None else v).strip().lstrip("'")


async def tra_cuu(nguon: dict, gia_tri: str) -> dict | None:
    """Tìm dòng có `lookup_key` khớp `gia_tri`. None nếu không có."""
    khoa = nguon.get("lookup_key", "")
    if not khoa:
        return None
    rows, loi = await doc_du_lieu(nguon)
    if loi:
        logger.warning(f"[tra cứu] {nguon.get('name')}: {loi}")
        return None

    can = _chuan_khoa(gia_tri)
    if not can:
        return None
    for r in rows:
        if _chuan_khoa(r.get(khoa, "")) == can:
            return r
    return None


async def tra_cuu_tat_ca(gia_tri: str) -> list[dict]:
    """Tra `gia_tri` qua mọi nguồn đang bật ở chế độ lookup.

    Dùng khi mở phiên: có sẵn thông tin của chính khách này thì bot tư vấn được
    theo hồ sơ thật thay vì nói chung chung.
    """
    ket_qua = []
    for nguon in await list_sources():
        if not nguon["enabled"] or nguon["mode"] != "lookup":
            continue
        try:
            dong = await tra_cuu(nguon, gia_tri)
        except Exception as e:
            logger.warning(f"[tra cứu] nguồn {nguon.get('name')} lỗi: {e}")
            continue
        if dong:
            ket_qua.append({"nguon": nguon["name"], "du_lieu": dong})
    return ket_qua


def dung_ngu_canh(ket_qua: list[dict]) -> str:
    """Biến kết quả tra cứu thành đoạn văn chèn vào THÔNG TIN THAM KHẢO."""
    if not ket_qua:
        return ""
    # MỖI TRƯỜNG MỘT DÒNG, tên trường viết bằng tiếng Việt có dấu.
    #
    # Bản cũ dồn tất cả vào một dòng ngăn bằng "|" và dùng nguyên tên cột của
    # CSDL ("du_no", "ky_han_con"). Đo trên máy thật: hồ sơ tra đúng, nằm đúng
    # trong prompt, mà mô hình 3B vẫn trả lời "em chưa biết số dư cụ thể" - nó
    # không nhặt được con số ra khỏi dòng dày đặc chữ không dấu.
    _TEN = {
        "so_dien_thoai": "Số điện thoại", "ho_ten": "Họ tên",
        "gioi_tinh": "Giới tính", "san_pham": "Sản phẩm đang dùng",
        "ma_hop_dong": "Mã hợp đồng", "du_no": "Dư nợ hiện tại",
        "ky_han_con": "Kỳ hạn còn lại", "ngay_den_han": "Ngày đến hạn",
        "tra_hang_thang": "Số tiền phải trả hàng tháng",
        "lai_suat_dang_ap_dung": "Lãi suất đang áp dụng cho khách này",
        "han_muc_da_duyet": "Hạn mức đã duyệt cho khách này",
        "nhom_no": "Nhóm nợ", "han_muc": "Hạn mức", "lai_suat": "Lãi suất",
    }
    # KHÔNG đưa vào prompt: chỉ để hệ thống dùng, nói ra là lộ chuyện theo dõi
    # khách ("anh đã từ chối 5 lần rồi") - phản tác dụng ngay trong cuộc gọi.
    # Bộ quay số và người trực dùng chúng qua Báo cáo, không qua lời AI.
    _KHONG_NOI = {"so_lan_da_goi", "ket_qua_lan_truoc", "ly_do_tu_choi",
                  "khung_gio_nen_goi", "ghi_chu"}
    dong = []
    for item in ket_qua:
        for k, v in item["du_lieu"].items():
            if k in _KHONG_NOI:
                continue
            if str(v).strip():
                dong.append(f"- {_TEN.get(k, k.replace('_', ' '))}: {v}")
    return "HỒ SƠ CỦA CHÍNH KHÁCH ĐANG NGHE MÁY:\n" + "\n".join(dong)
