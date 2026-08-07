"""API nguồn dữ liệu ngoài (Điều 6: kết nối CSDL/Excel để tra cứu tư vấn)."""

import logging
import re
import sqlite3
import time
from pathlib import Path

from fastapi import APIRouter, File, UploadFile
from pydantic import BaseModel

from backend.services import data_source_service as ds

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/data-sources", tags=["data-sources"])

# File tải lên nằm gọn một chỗ, không rải khắp máy: người dùng còn phải sao lưu
# và biết cái gì đang được hệ thống đọc.
THU_MUC = Path("./data/nguon")
_DUOI_SQLITE = {".db", ".sqlite", ".sqlite3", ".db3"}
_DUOI_NHAN = _DUOI_SQLITE | {".csv", ".xlsx", ".xls"}
_TRAN_UPLOAD = 200 * 1024 * 1024


class NguonRequest(BaseModel):
    source_id: str = ""
    name: str = "Nguồn dữ liệu"
    kind: str = "excel"              # excel | csv | sqlite
    mode: str = "rag"                # rag | lookup
    path: str = ""                   # đường dẫn file cục bộ
    sheet: str = ""                  # excel: tên sheet, trống = sheet đầu
    table: str = ""                  # sqlite: tên bảng
    query: str = ""                  # sqlite: câu SELECT thay cho tên bảng
    content_columns: list[str] = []  # rag: chỉ nạp các cột này, trống = tất cả
    lookup_key: str = ""             # lookup: cột khoá, vd 'so_dien_thoai'
    knowledge_tag: str = ""
    enabled: bool = True


class TraCuuRequest(BaseModel):
    gia_tri: str


def _config(req: NguonRequest) -> dict:
    return {
        "path": req.path.strip(),
        "sheet": req.sheet.strip(),
        "table": req.table.strip(),
        "query": req.query.strip(),
        "content_columns": [c for c in req.content_columns if c.strip()],
    }


@router.post("/upload")
async def upload_file(file: UploadFile = File(...)):
    """Nhận file rồi TỰ DÒ cấu trúc bên trong.

    Trước đây trang này chỉ có ô gõ đường dẫn. Muốn dùng thì phải tự copy file
    vào máy chủ, tự biết đường dẫn tuyệt đối, tự biết tên bảng trong file .db và
    tên cột khoá — gõ sai một chỗ là báo lỗi chung chung. Người vận hành bán hàng
    không làm được việc đó.

    Nên trả luôn danh sách bảng/sheet kèm số dòng và tên cột: người dùng CHỌN chứ
    không gõ.
    """
    ten = Path(file.filename or "nguon").name
    duoi = Path(ten).suffix.lower()
    if duoi not in _DUOI_NHAN:
        return {"error": f"Chưa nhận đuôi '{duoi}'. Nhận: {', '.join(sorted(_DUOI_NHAN))}"}

    raw = await file.read()
    if not raw:
        return {"error": "File rỗng"}
    if len(raw) > _TRAN_UPLOAD:
        return {"error": f"File {len(raw)//1024//1024} MB, vượt trần "
                         f"{_TRAN_UPLOAD//1024//1024} MB"}

    # Tên file giữ nguyên phần gốc cho dễ nhận, nhưng lọc sạch ký tự lạ: giá trị
    # này thành đường dẫn thật trên đĩa.
    goc = re.sub(r"[^0-9A-Za-z._-]+", "_", Path(ten).stem)[:60] or "nguon"
    THU_MUC.mkdir(parents=True, exist_ok=True)
    dich = THU_MUC / f"{goc}{duoi}"
    if dich.exists():
        dich = THU_MUC / f"{goc}_{int(time.time())}{duoi}"
    dich.write_bytes(raw)
    logger.info("Nguồn dữ liệu: nhận %s (%d KB)", dich, len(raw) // 1024)

    return {
        "ok": True,
        "path": str(dich.resolve()),
        "ten_tep": dich.name,
        "kich_thuoc": len(raw),
        "kind": "sqlite" if duoi in _DUOI_SQLITE else ("csv" if duoi == ".csv" else "excel"),
        "cau_truc": _do_cau_truc(dich, duoi),
    }


def _do_cau_truc(p: Path, duoi: str) -> dict:
    """Liệt kê bảng/sheet + cột + số dòng. Lỗi thì báo, đừng ném ra ngoài."""
    try:
        if duoi in _DUOI_SQLITE:
            # mode=ro: chỉ đọc. File người dùng vừa tải lên có thể là CSDL đang
            # được ứng dụng khác dùng - mở ghi là rủi ro không cần thiết.
            con = sqlite3.connect(f"file:{p.as_posix()}?mode=ro", uri=True)
            bang = []
            for (t,) in con.execute(
                "SELECT name FROM sqlite_master WHERE type IN ('table','view') "
                "AND name NOT LIKE 'sqlite_%' ORDER BY name"
            ):
                try:
                    n = con.execute(f'SELECT COUNT(*) FROM "{t}"').fetchone()[0]
                except Exception:
                    n = None
                cot = [r[1] for r in con.execute(f'PRAGMA table_info("{t}")')]
                bang.append({"ten": t, "so_dong": n, "cot": cot})
            con.close()
            return {"loai": "sqlite", "bang": bang}

        if duoi == ".csv":
            import csv as _csv
            with p.open("r", encoding="utf-8-sig", errors="replace", newline="") as f:
                doc = _csv.reader(f)
                cot = next(doc, [])
                n = sum(1 for _ in doc)
            return {"loai": "csv", "bang": [{"ten": p.name, "so_dong": n, "cot": cot}]}

        from openpyxl import load_workbook
        # read_only + values_only: bảng lãi suất vài chục nghìn dòng nạp cả công
        # thức vào RAM chỉ để đếm dòng là phí.
        wb = load_workbook(p, read_only=True, data_only=True)
        bang = []
        for ws in wb.worksheets:
            dong_dau = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
            bang.append({
                "ten": ws.title,
                "so_dong": max((ws.max_row or 1) - 1, 0),
                "cot": [str(c) for c in dong_dau if c is not None],
            })
        wb.close()
        return {"loai": "excel", "bang": bang}
    except Exception as e:
        return {"loai": duoi.lstrip("."), "bang": [], "loi": f"Không dò được cấu trúc: {e}"}


@router.get("")
async def list_sources():
    return {
        "sources": await ds.list_sources(),
        "kinds": ds.KINDS,
        "modes": ds.MODES,
        "giai_thich": {
            "rag": "Tri thức chung (bảng giá, biểu phí) — bot tìm gần đúng theo nghĩa",
            "lookup": "Tra theo khách (dư nợ, đơn hàng) — khớp chính xác theo cột khoá",
        },
    }


@router.post("")
async def upsert_source(req: NguonRequest):
    if req.mode == "lookup" and not req.lookup_key.strip():
        return {"error": "Chế độ tra cứu phải khai cột khoá (ví dụ: so_dien_thoai)"}
    nguon = await ds.upsert_source(
        req.source_id,
        name=req.name.strip() or "Nguồn dữ liệu",
        kind=req.kind,
        mode=req.mode,
        config=_config(req),
        lookup_key=req.lookup_key.strip(),
        knowledge_tag=req.knowledge_tag.strip(),
        enabled=req.enabled,
    )
    if nguon is None:
        return {"error": "Không lưu được — database chưa sẵn sàng"}
    return {"source": nguon}


@router.delete("/{source_id}")
async def delete_source(source_id: str):
    """Xoá nguồn và gỡ luôn dữ liệu nó đã nạp vào kho tri thức."""
    from backend.main import app_state

    if not await ds.delete_source(source_id, rag=app_state.rag):
        return {"error": "Nguồn không tồn tại"}
    return {"deleted": source_id}


@router.post("/{source_id}/preview")
async def preview_source(source_id: str, limit: int = 10):
    """Xem trước 10 dòng đầu và danh sách cột, trước khi nạp."""
    nguon = await ds.get_source(source_id)
    if nguon is None:
        return {"error": "Nguồn không tồn tại"}

    rows, loi = await ds.doc_du_lieu(nguon)
    if loi:
        return {"error": loi}
    return {
        "so_dong": len(rows),
        "cot": list(rows[0].keys()) if rows else [],
        "xem_truoc": rows[:limit],
        "vi_du_van_ban": ds._dong_thanh_van_ban(
            rows[0], (nguon.get("config") or {}).get("content_columns") or []
        ) if rows else "",
    }


@router.post("/{source_id}/sync")
async def sync_source(source_id: str):
    """Nạp lại vào kho tri thức. Chỉ áp dụng cho nguồn chế độ 'rag'."""
    from backend.main import app_state

    nguon = await ds.get_source(source_id)
    if nguon is None:
        return {"error": "Nguồn không tồn tại"}
    if nguon["mode"] != "rag":
        return {
            "error": "Nguồn chế độ tra cứu không cần nạp — nó đọc thẳng từ file "
                     "mỗi lần tra, nên sửa file là có hiệu lực ngay"
        }
    ket_qua = await ds.nap_vao_rag(nguon, app_state.rag)
    if ket_qua.get("error"):
        return ket_qua
    return {**ket_qua, "source": await ds.get_source(source_id)}


@router.post("/{source_id}/lookup")
async def lookup(source_id: str, req: TraCuuRequest):
    """Tra thử một giá trị khoá, để kiểm cấu hình trước khi đem dùng thật."""
    nguon = await ds.get_source(source_id)
    if nguon is None:
        return {"error": "Nguồn không tồn tại"}
    if nguon["mode"] != "lookup":
        return {"error": "Nguồn này ở chế độ 'rag', không tra theo khoá được"}

    dong = await ds.tra_cuu(nguon, req.gia_tri)
    if dong is None:
        return {"tim_thay": False, "ghi_chu": f"Không có dòng nào khớp '{req.gia_tri}'"}
    return {"tim_thay": True, "du_lieu": dong}


@router.post("/lookup-all")
async def lookup_all(req: TraCuuRequest):
    """Tra một giá trị qua mọi nguồn tra cứu đang bật."""
    ket_qua = await ds.tra_cuu_tat_ca(req.gia_tri)
    return {"ket_qua": ket_qua, "ngu_canh": ds.dung_ngu_canh(ket_qua)}
